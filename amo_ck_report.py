import json
import random
import time
from collections import Counter
from typing import Dict, Optional

import requests
from openpyxl import Workbook, load_workbook

import config


def log(message: str) -> None:
    if not config.AMO_DEBUG:
        return
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {message}")


class AmoClient:
    def __init__(self, base_url: str, tokens_file: str):
        self.base_url = base_url.rstrip("/")
        self.tokens_file = tokens_file
        self.tokens = self._load_tokens()

    def _load_tokens(self) -> Dict[str, str]:
        try:
            with open(self.tokens_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            return {}

    def _save_tokens(self, tokens: Dict[str, str]) -> None:
        with open(self.tokens_file, "w", encoding="utf-8") as f:
            json.dump(tokens, f, ensure_ascii=False, indent=2)

    def _mask(self, token: str) -> str:
        if not token:
            return ""
        if len(token) <= 10:
            return token[:2] + "***"
        return f"{token[:6]}***{token[-4:]}"

    def _refresh_tokens(self) -> None:
        refresh_token = self.tokens.get("refresh_token")
        if not refresh_token:
            raise RuntimeError("refresh_token is missing in tokens.json")
        url = f"{self.base_url}/oauth2/access_token"
        payload = {
            "client_id": config.CLIENT_ID,
            "client_secret": config.CLIENT_SECRET,
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "redirect_uri": config.REDIRECT_URI,
        }
        resp = requests.post(url, json=payload, timeout=20)
        log(f"POST {url} -> {resp.status_code}")
        resp.raise_for_status()
        data = resp.json()
        self.tokens = {
            "access_token": data.get("access_token"),
            "refresh_token": data.get("refresh_token"),
            "updated_at": int(time.time()),
        }
        self._save_tokens(self.tokens)
        log(
            f"Tokens refreshed: access={self._mask(self.tokens.get('access_token'))}, "
            f"refresh={self._mask(self.tokens.get('refresh_token'))}"
        )

    def request(self, method: str, path: str, params=None, retry=0):
        url = f"{self.base_url}{path}"
        headers = {
            "Authorization": f"Bearer {self.tokens.get('access_token', '')}",
            "Accept": "application/hal+json",
        }
        try:
            log(f"{method} {url} params={params} attempt={retry + 1}")
            resp = requests.request(method, url, headers=headers, params=params, timeout=30)
            log(f"{method} {url} -> {resp.status_code}")
        except requests.RequestException:
            if retry >= 3:
                raise
            time.sleep(1 + retry)
            return self.request(method, path, params=params, retry=retry + 1)

        if resp.status_code == 401:
            if retry >= 1:
                resp.raise_for_status()
            self._refresh_tokens()
            return self.request(method, path, params=params, retry=retry + 1)

        if resp.status_code == 429:
            if retry >= 5:
                resp.raise_for_status()
            time.sleep(random.uniform(1, 3))
            return self.request(method, path, params=params, retry=retry + 1)

        resp.raise_for_status()
        return resp


def load_users_map(client: AmoClient) -> Dict[str, str]:
    users = {}
    try:
        page = 1
        while True:
            resp = client.request("GET", "/api/v4/users", params={"limit": 250, "page": page, "with": "role,group,phone_number"})
            data = resp.json()
            items = data.get("_embedded", {}).get("users", [])
            log(f"Users page {page}: {len(items)}")
            if not items:
                break
            for user in items:
                uid = str(user.get("id"))
                name = user.get("name") or f"User {uid}"
                users[uid] = name
            page += 1
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 403:
            print("Users API returned 403, falling back to users_map.json")
            return load_users_map_file()
        raise
    return users


def load_users_map_file() -> Dict[str, str]:
    try:
        with open(config.USERS_MAP_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return {str(k): v for k, v in data.items()}
    except FileNotFoundError:
        return {}


def is_ck_lead(lead) -> bool:
    if lead.get("is_deleted"):
        return False
    fields = lead.get("custom_fields_values") or []
    for field in fields:
        if int(field.get("field_id", 0)) != config.CK_FIELD_ID:
            continue
        for val in field.get("values") or []:
            if int(val.get("enum_id", 0)) == config.CK_ENUM_ID:
                return True
            if val.get("value") == "ЦК":
                return True
    return False


def iter_leads(client: AmoClient):
    page = 1
    processed = 0
    while True:
        resp = client.request("GET", "/api/v4/leads", params={"limit": 250, "page": page})
        data = resp.json()
        leads = data.get("_embedded", {}).get("leads", [])
        log(f"Leads page {page}: {len(leads)}")
        if not leads:
            break
        for lead in leads:
            processed += 1
            yield lead
        page += 1
    return processed


def ensure_headers(sheet, headers):
    header_row = 1
    current = {cell.value: idx + 1 for idx, cell in enumerate(sheet[header_row]) if cell.value}
    for name in headers:
        if name not in current:
            col = len(current) + 1
            sheet.cell(row=header_row, column=col, value=name)
            current[name] = col
    return current


def write_counts_to_excel(users_map, counts_responsible):
    try:
        wb = load_workbook(config.EXCEL_PATH)
    except FileNotFoundError:
        wb = Workbook()

    if config.EXCEL_SHEET in wb.sheetnames:
        sheet = wb[config.EXCEL_SHEET]
    else:
        sheet = wb.active
        sheet.title = config.EXCEL_SHEET

    headers = [
        config.EMPLOYEE_ID_COLUMN,
        config.EMPLOYEE_NAME_COLUMN,
        config.COL_CK_OPERATOR,
    ]
    header_map = ensure_headers(sheet, headers)

    all_ids = set(counts_responsible)
    rows_by_id = {}
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=header_map[config.EMPLOYEE_ID_COLUMN]).value
        if cell_value is not None:
            rows_by_id[str(cell_value)] = row

    current_row = sheet.max_row + 1
    for uid in sorted(all_ids, key=lambda x: int(x) if x.isdigit() else x):
        row = rows_by_id.get(uid)
        if not row:
            row = current_row
            current_row += 1
        sheet.cell(row=row, column=header_map[config.EMPLOYEE_ID_COLUMN], value=uid)
        sheet.cell(
            row=row,
            column=header_map[config.EMPLOYEE_NAME_COLUMN],
            value=users_map.get(uid, f"User {uid}")
        )
        sheet.cell(row=row, column=header_map[config.COL_CK_OPERATOR], value=counts_responsible.get(uid, 0))

    wb.save(config.EXCEL_PATH)


def print_top(title: str, counter: Counter):
    if not counter:
        print(f"{title}: нет данных")
        return
    top = counter.most_common(5)
    formatted = ", ".join([f"{uid}={count}" for uid, count in top])
    print(f"{title}: {formatted}")


def main():
    client = AmoClient(config.AMO_BASE_URL, config.TOKENS_FILE)
    users_map = load_users_map(client)
    if not users_map:
        users_map = load_users_map_file()
        log(f"Users map fallback size: {len(users_map)}")

    ck_total = 0
    ck_responsible = Counter()

    pages = 0
    processed = 0

    for lead in iter_leads(client):
        processed += 1
        if processed % 250 == 1:
            pages += 1
        if not lead or lead.get("is_deleted"):
            continue

        if not is_ck_lead(lead):
            continue

        ck_total += 1
        responsible_id = str(lead.get("responsible_user_id") or "")
        if responsible_id:
            ck_responsible[responsible_id] += 1
        log(f"CK lead id={lead.get('id')} responsible={responsible_id}")

    print(f"Pages fetched: {pages}")
    print(f"Leads processed: {processed}")
    print(f"CK total: {ck_total}")
    print_top("Top responsible", ck_responsible)

    write_counts_to_excel(users_map, ck_responsible)


if __name__ == "__main__":
    main()
